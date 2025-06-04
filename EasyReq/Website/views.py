import json, re
from django.db.models import Count, Q, Avg
from django.http import HttpResponseRedirect, HttpResponse
from django.contrib.auth.forms import AuthenticationForm, PasswordChangeForm
from django.contrib.auth import login, logout, authenticate, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.views.generic import CreateView
from openai import OpenAI
from .models import User, Request, RequestStatus, RequestComment, Department, Review
from .forms import RegistrationForm
from django.http import JsonResponse
from django.db import models
from django.core.paginator import Paginator
from .utils import send_request_notification
from django.utils import timezone
import csv
import openpyxl
from openpyxl.styles import Font, Alignment
import io
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.conf import settings
from django.views.decorators.http import require_POST


from .models import Notification
import logging
logger = logging.getLogger(__name__)

def home(request):
    if request.user.is_authenticated and request.user.role in [1,2,3]:
        return redirect('dashboard')

    recent_requests = []
    student_status_counts = [0,0,0]
    if request.user.is_authenticated and request.user.role == 0:
        recent_requests = (
            Request.objects
                   .filter(student=request.user)
                   .order_by('-created')[:5]
        )
        qs = (
            Request.objects
                   .filter(student=request.user)
                   .values('status')
                   .annotate(count=Count('status'))
        )
        m = {item['status']:item['count'] for item in qs}
        student_status_counts = [m.get(0,0), m.get(1,0), m.get(2,0)]

    return render(request, 'home.html', {
        'recent_requests': recent_requests,
        'student_status_counts': student_status_counts,
    })


@require_POST
@login_required
def bulk_delete_requests(request):
    user = request.user
    request_ids = request.POST.getlist('selected_requests')  

    if not request_ids:
        messages.warning(request, "לא נבחרו בקשות למחיקה.")
        return redirect('list_requests')

    deleted_count = 0
    for req_id in request_ids:
        req = get_object_or_404(Request, id=req_id)
        if user == req.student or user.role in [1, 2, 3]:
            req.delete()
            deleted_count += 1

    messages.success(request, f"{deleted_count} בקשות נמחקו בהצלחה.")
    return redirect('list_requests')



@login_required
def bulk_delete_students(request):
    if request.method == 'POST' and request.POST.get('bulk_delete'):
        selected_students = request.POST.getlist('selected_students')
        if selected_students:
            # מחיקת הסטודנטים
            User.objects.filter(id__in=selected_students, role=0).delete()
            messages.success(request, f'{len(selected_students)} סטודנטים נמחקו בהצלחה')
        else:
            messages.error(request, 'לא נבחרו סטודנטים למחיקה')
    
    return redirect('manage_users')

@login_required
def bulk_deactivate_lecturers(request):
    if request.method == 'POST' and request.POST.get('bulk_deactivate'):
        selected_lecturers = request.POST.getlist('selected_lecturers')
        if selected_lecturers:
            # השבתת המרצים
            User.objects.filter(id__in=selected_lecturers, role=2, is_active=True).update(is_active=False)
            messages.success(request, f'{len(selected_lecturers)} מרצים הושבתו בהצלחה')
        else:
            messages.error(request, 'לא נבחרו מרצים להשבתה')
    
    return redirect('manage_users')


@require_POST
@login_required
def bulk_delete_courses(request):
    if request.user.role != 2:
        messages.error(request, "אין לך הרשאות לבצע פעולה זו.")
        return redirect('home')

    selected_courses = request.POST.getlist('selected_courses')
    if selected_courses:
        Course.objects.filter(id__in=selected_courses, dept=request.user.department).delete()
        messages.success(request, f'{len(selected_courses)} קורסים נמחקו בהצלחה')
    else:
        messages.error(request, "לא נבחרו קורסים למחיקה.")

    return redirect('manage_courses')


def register_view(request):
    if request.method == "POST":
        form = RegistrationForm(request.POST, request.FILES)
        if form.is_valid():
            user = form.save(commit=False)
            user.save()
            if user.role == 1: #lecturer 
                user.is_active = 0
                selected_courses = request.POST.getlist('courses')
                if selected_courses:
                    for course_id in selected_courses:
                        user.courses.add(course_id)

                # send notification to secretary
                staff_users = User.objects.filter(role=2, department=user.department)
                for staff_user in staff_users:
                    # Create an alert for new lecturer registration
                    Notification.objects.create(
                        user=staff_user,
                        message=f"מרצה חדש ({user.get_full_name()}) נרשם וממתין לאישור",
                        read=False
                    )

            if user.role == 0: 
                relevant_courses = Course.objects.filter(year=user.year, dept=user.department)
                if relevant_courses.exists():
                    user.courses.add(*relevant_courses)
            return redirect('registration_success', user_id=user.id)
    else:
        form = RegistrationForm()

    departments = Department.objects.all()
    courses = Course.objects.all()
    return render(request, 'register.html', {"form": form, "departments": departments, "courses": courses})

def registration_success(request, user_id):
    try:
        user = User.objects.get(id=user_id)
        role = user.role
        context = {'user': user, 'role': role}
        return render(request, 'registration_success.html', context)
    except User.DoesNotExist:
        return redirect('register')


def get_courses(request):
    department_id = request.GET.get('department')
    if not department_id:
        return JsonResponse({'courses': []})
    courses = Course.objects.filter(dept=department_id).values('id', 'name')
    return JsonResponse({'courses': list(courses)})

def login_view(request):
    if request.user.is_authenticated:
        return redirect("/Website/home")
    return render(request, 'login.html')

def login_request(request):
    if request.user.is_authenticated:
        return redirect('home')

    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        username = form.data.get('username')
        try:
            user_obj = User.objects.get(username=username)
            if not user_obj.is_active:
                messages.error(request, "המשתמש שלך ממתין לאישור מנהל מערכת")
                return HttpResponseRedirect("/login")
        except User.DoesNotExist:
            pass

        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('home')
            else:
                messages.error(request, "שגיאה באחד מפרטי הזיהוי")
        else:
            messages.error(request, "שגיאה באחד מפרטי הזיהוי")

    return render(request, 'login.html', {'form': AuthenticationForm()})

@login_required
def logout_view(request):
    logout(request)
    return redirect("home")

class Registration(CreateView):
    model = User
    form_class = RegistrationForm
    template_name = 'register.html'

    def form_valid(self, form):
        user = form.save()
        path_ = "/registration-success/" + str(user.id)
        return redirect(path_, user_id=user.id)


@login_required
def profile_view(request):
    user = request.user

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'update_email' and user.role == 0:  # Students only
            new_email = request.POST.get('new_email')

            # Validate email format
            pattern = r'^[\w\.-]+@(ac\.sce\.ac\.il|sce\.ac\.il)$'
            if not re.match(pattern, new_email):
                messages.error(request, "עליך להזין מייל מכללתי בלבד")
                return redirect('profile')

            # Check if email is already in use
            if User.objects.filter(email=new_email).exclude(id=user.id).exists():
                messages.error(request, "כתובת האימייל כבר בשימוש")
                return redirect('profile')

            # Update email immediately
            user.email = new_email
            user.save()
            messages.success(request, "כתובת האימייל עודכנה בהצלחה!")

        elif action == 'update_courses' and user.role == 1:  # Lecturers only
            selected_course_ids = request.POST.getlist('selected_courses')

            # Clear current courses and add selected ones
            user.courses.clear()
            if selected_course_ids:
                valid_courses = Course.objects.filter(
                    id__in=selected_course_ids,
                    dept=user.department
                )
                user.courses.add(*valid_courses)

            messages.success(request, "רשימת הקורסים עודכנה בהצלחה!")

        elif action == 'update_password':
            password_form = PasswordChangeForm(user, request.POST)
            if password_form.is_valid():
                user = password_form.save()
                update_session_auth_hash(request, user)
                messages.success(request, "הסיסמה עודכנה בהצלחה!")
            else:
                for field, errors in password_form.errors.items():
                    for error in errors:
                        messages.error(request, f"{error}")

        elif 'profile_pic' in request.FILES:
            user.profile_pic = request.FILES['profile_pic']
            user.save()
            messages.success(request, "תמונת הפרופיל עודכנה בהצלחה!")

        return redirect('profile')

    else:
        password_form = PasswordChangeForm(user)

    # Get all courses in department for lecturers
    all_department_courses = []
    if user.role == 1 and user.department:
        all_department_courses = Course.objects.filter(dept=user.department).order_by('year', 'name')

    return render(request, 'profile.html', {
        'user': user,
        'password_form': password_form,
        'all_department_courses': all_department_courses,
    })

@login_required
def create_request(request):
    if request.method == 'POST':
        student = request.user
        dept_id = student.department_id
        course_id = request.POST.get('course')
        title_id = request.POST.get('title')
        description = request.POST.get('description')
        priority = request.POST.get('priority', 1)

        # Optional: handle file upload
        attachment = request.FILES.get('attachment')
        new_request = Request(student=student,dept_id=dept_id,title=int(title_id),description=description,
                              priority=priority,attachments=attachment)

        if course_id:
            new_request.course_id = course_id

        new_request.save()
        assigned_users = new_request.auto_assign()
        RequestStatus.objects.create(request=new_request,status=0, updated_by=request.user,notes="Request submitted")

        if len(assigned_users) > 0:
            send_request_notification(new_request, 'created')
            for user in assigned_users:
                send_request_notification(new_request, 'assigned', user)

        return redirect('request_detail', request_id=new_request.id)

    departments = Department.objects.all()
    courses = request.user.courses.all() if request.user.role == 0 else []

    # Get available templates
   # if request.user.role == 0:  # Only students see templates
   #     templates = RequestTemplate.objects.filter(
   #         dept=request.user.department,
   #         is_active=True
   #     )
   # else:
    #    templates = []

    return render(request, 'create_request.html', {'departments': departments,'courses': courses,
                'title_choices': Request.TITLES,})
    #'templates': templates
import os
from django.shortcuts import get_object_or_404, redirect


def request_detail(request, request_id):
    req = get_object_or_404(Request, pk=request_id)
    assigned_users = list(req.assigned_to.all())
    logger.debug(f"Request #{req.id} is assigned to: {assigned_users}")

    has_access = (
        request.user == req.student or
        req.assigned_to.filter(id=request.user.id).exists() or
        request.user.role in [2, 3] or
        req.viewers.filter(id=request.user.id).exists()
    )

    if not has_access:
        messages.error(request, "אין לך הרשאות לצפות בבקשה זו.")
        return redirect('list_requests')

    if request.method == 'POST':
        comment = request.POST.get('comment')
        if comment:
            RequestComment.objects.create(request=req, user=request.user, comment=comment)
            if request.user == req.student:
                for assigned_user in req.assigned_to.all():
                    send_request_notification(req, 'comment', assigned_user)
                    logger.debug(f"Sent comment notification to assigned user: {assigned_user}")
            else:
                send_request_notification(req, 'comment', req.student)
                logger.debug(f"Sent comment notification to student: {req.student}")
        return redirect('request_detail', request_id=req.id)

    comments = req.comments.all()

    file_url = req.attachments.url if req.attachments else None
    file_type = None
    if file_url:
        ext = os.path.splitext(file_url)[1].lower()
        if ext == ".pdf":
            file_type = "pdf"
        elif ext in [".jpg", ".jpeg", ".png", ".gif"]:
            file_type = "image"

    return render(request, 'request_detail.html', {
        'request': req,
        'comments': comments,
        'sla_status': req.get_sla_status(),
        'file_url': file_url,
        'file_type': file_type
    })

def update_request_status(request, request_id):
    PIPELINE_STATUSES = (
        (0, 'נשלח'),
        (1, 'בבדיקה'),
        (2, 'בטיפול'),
        (3, 'בהמתנה למסמכים נוספים'),
        (4, 'טופל - אושר'),
        (5, 'טופל - נדחה'),
        (6, 'בהמתנה')
    )

    if not request.user.role in [1, 2, 3]:  # staff
        return redirect('home')

    req = Request.objects.get(pk=request_id)

    if request.method == 'POST':
        new_pipeline_status = int(request.POST.get('pipeline_status'))
        status_notes = request.POST.get('status_notes', '')
        resolution_notes = request.POST.get('resolution_notes', '')
        req.pipeline_status = new_pipeline_status
        if new_pipeline_status in [4, 5]:  # Resolved
            # 4=approved, 5=rejected
            req.status = 1 if new_pipeline_status == 4 else 2
            req.resolution_notes = resolution_notes
            req.resolved_date = timezone.now()
        else:
            req.status = 0

        req.save()
        RequestStatus.objects.create(request=req,status=new_pipeline_status,updated_by=request.user,notes=status_notes)
        pipeline_status_text = dict(PIPELINE_STATUSES)[new_pipeline_status]
        RequestComment.objects.create(request=req,user=request.user,
            comment=f"סטטוס עודכן ל: {pipeline_status_text}\n\n{status_notes}")

        notification_type = None

        if new_pipeline_status in [4, 5]:
            notification_type = 'resolved'
        elif new_pipeline_status == 3:
            notification_type = 'info_requested'
        else:
            notification_type = 'updated'

        if 'notify_student' in request.POST:
            send_request_notification(req, notification_type)
            if notification_type == 'resolved' and req.assigned_to.exists():
                for assigned_user in req.assigned_to.all():
                    send_request_notification(req, 'resolved', assigned_user)

        return redirect('request_detail', request_id=req.id)

    return render(request, 'update_request.html', {'request': req,'statuses': PIPELINE_STATUSES})
from django.views.decorators.http import require_http_methods
from django.shortcuts import get_object_or_404

@require_http_methods(["GET", "POST"])
@login_required
def list_requests(request):
    user = request.user

    if request.method == "POST" and "delete_request_id" in request.POST:
        request_id = request.POST.get("delete_request_id")
        req_to_delete = get_object_or_404(Request, id=request_id)

        if user == req_to_delete.student or user.role in [1, 2, 3]:
            req_to_delete.delete()
            messages.success(request, "הבקשה נמחקה בהצלחה.")
            return redirect('list_requests')
        else:
            messages.error(request, "אין לך הרשאה למחוק בקשה זו.")
            return redirect('list_requests')

    if user.role == 0:
        base_queryset = Request.objects.filter(student=user)
    elif user.role in [1, 2, 3]:
        base_queryset = Request.objects.filter(
            models.Q(viewers=user) | models.Q(assigned_to=user)
        ).distinct()
    else:
        base_queryset = Request.objects.none()

    search_query = request.GET.get('q')
    if search_query:
        base_queryset = base_queryset.filter(
            models.Q(title__icontains=search_query) |
            models.Q(description__icontains=search_query) |
            models.Q(resolution_notes__icontains=search_query)
        )

    status = request.GET.get('status')
    if status and status.isdigit():
        base_queryset = base_queryset.filter(status=int(status))

    pipeline_status = request.GET.get('pipeline_status')
    if pipeline_status and pipeline_status.isdigit():
        base_queryset = base_queryset.filter(pipeline_status=int(pipeline_status))

    priority = request.GET.get('priority')
    if priority and priority.isdigit():
        base_queryset = base_queryset.filter(priority=int(priority))

    department = request.GET.get('department')
    if department and department.isdigit():
        base_queryset = base_queryset.filter(dept_id=int(department))

    # Request title filter (חדש)
    request_title = request.GET.get('request_title')
    if request_title and request_title.isdigit():
        base_queryset = base_queryset.filter(title=int(request_title))

    # Date range filter (חדש)
    date_range = request.GET.get('date_range')
    if date_range:
        from datetime import datetime, timedelta
        today = timezone.now().date()
        
        if date_range == 'today':
            base_queryset = base_queryset.filter(created__date=today)
        elif date_range == 'yesterday':
            yesterday = today - timedelta(days=1)
            base_queryset = base_queryset.filter(created__date=yesterday)
        elif date_range == 'this_week':
            start_week = today - timedelta(days=today.weekday())
            base_queryset = base_queryset.filter(created__date__gte=start_week)
        elif date_range == 'last_week':
            start_week = today - timedelta(days=today.weekday() + 7)
            end_week = today - timedelta(days=today.weekday() + 1)
            base_queryset = base_queryset.filter(created__date__range=[start_week, end_week])
        elif date_range == 'this_month':
            start_month = today.replace(day=1)
            base_queryset = base_queryset.filter(created__date__gte=start_month)
        elif date_range == 'last_month':
            first_day_this_month = today.replace(day=1)
            last_day_last_month = first_day_this_month - timedelta(days=1)
            first_day_last_month = last_day_last_month.replace(day=1)
            base_queryset = base_queryset.filter(created__date__range=[first_day_last_month, last_day_last_month])
        elif date_range == 'this_year':
            start_year = today.replace(month=1, day=1)
            base_queryset = base_queryset.filter(created__date__gte=start_year)

    sla = request.GET.get('sla')
    if sla:
        pending = base_queryset.filter(status=0)
        if sla == 'overdue':
            overdue_ids = [req.id for req in pending if req.get_sla_status() == "בחריגה"]
            base_queryset = base_queryset.filter(id__in=overdue_ids)
        elif sla == 'at_risk':
            at_risk_ids = [req.id for req in pending if req.get_sla_status() == "בסיכון"]
            base_queryset = base_queryset.filter(id__in=at_risk_ids)
        elif sla == 'on_track':
            on_track_ids = [req.id for req in pending if req.get_sla_status() == "בזמן"]
            base_queryset = base_queryset.filter(id__in=on_track_ids)

    # Custom date range (נשאר כמו שהיה)
    from datetime import datetime
    date_from = request.GET.get('date_from')
    if date_from:
        try:
            date_from = datetime.strptime(date_from, '%Y-%m-%d')
            base_queryset = base_queryset.filter(created__gte=date_from)
        except ValueError:
            pass

    date_to = request.GET.get('date_to')
    if date_to:
        try:
            date_to = datetime.strptime(date_to, '%Y-%m-%d')
            date_to = date_to.replace(hour=23, minute=59, second=59)
            base_queryset = base_queryset.filter(created__lte=date_to)
        except ValueError:
            pass

    # Order by created date (הוסרתי את המיון לפי title)
    base_queryset = base_queryset.order_by('-created')

    paginator = Paginator(base_queryset, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # ✅ העברת status_counts לכל המשתמשים (לא רק לסטודנטים)
    status_counts = [
        base_queryset.filter(status=0).count(),  # ממתין
        base_queryset.filter(status=1).count(),  # מאושר
        base_queryset.filter(status=2).count(),  # נדחה
    ]

    departments = Department.objects.all()

    # Get user-specific request titles (חדש)
    if user.role == 0:  # Student
        user_request_titles = Request.objects.filter(student=user).values_list('title', flat=True).distinct()
        user_request_titles = [(title, dict(Request.TITLES)[title]) for title in user_request_titles]
    elif user.role in [1, 2, 3]:  # Staff/Lecturers
        user_request_titles = Request.objects.filter(
            models.Q(viewers=user) | models.Q(assigned_to=user)
        ).values_list('title', flat=True).distinct()
        user_request_titles = [(title, dict(Request.TITLES)[title]) for title in user_request_titles]
    else:
        user_request_titles = []

    context = {
        'page_obj': page_obj,
        'departments': departments,
        'statuses': Request.STATUSES,
        'pipeline_statuses': Request.PIPELINE_STATUSES,
        'priorities': Request.PRIORITY_LEVELS,
        'user_request_titles': user_request_titles, 
        'search_query': search_query,
        'filters': {
            'status': request.GET.get('status'),
            'pipeline_status': request.GET.get('pipeline_status'),
            'priority': request.GET.get('priority'),
            'department': request.GET.get('department'),
            'request_title': request.GET.get('request_title'),  
            'date_range': request.GET.get('date_range'),  
            'date_from': request.GET.get('date_from'),
            'date_to': request.GET.get('date_to'),
            'sla': request.GET.get('sla'),
        },
        'status_counts': status_counts,  
    }

    return render(request, 'list_requests.html', context)

# notifications

@login_required
@require_http_methods(["POST"])
def mark_notification_read(request):
    """סימון התראה בודדת כנקראה"""
    try:
        data = json.loads(request.body)
        notification_id = data.get('notification_id')
        
        notification = get_object_or_404(Notification, id=notification_id, user=request.user)
        notification.read = True
        notification.save()
        
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
@require_http_methods(["POST"])
def mark_all_notifications_read(request):
    """סימון כל ההתראות כנקראות"""
    try:
        Notification.objects.filter(user=request.user, read=False).update(read=True)
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def all_notifications(request):
    """צפייה בכל ההתראות כולל אלו שנקראו"""
    notifications = Notification.objects.filter(user=request.user).order_by('-created_at')
    
    return render(request, 'all_notifications.html', {
        'notifications': notifications
    })

@login_required
@require_http_methods(["POST"])
def toggle_notification_status(request):
    """החלפת סטטוס התראה בין נקרא ללא נקרא"""
    try:
        data = json.loads(request.body)
        notification_id = data.get('notification_id')
        read_status = data.get('read') 
        
        notification = get_object_or_404(Notification, id=notification_id, user=request.user)
        notification.read = read_status
        notification.save()
        
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required()
def dashboard(request):
    """Dashboard view with widgets for staff and students with filtering support"""

    user = request.user
    today = timezone.now()
    thirty_days_ago = today - timezone.timedelta(days=30)
    
    if user.role == 0:  # Student
        base_queryset = Request.objects.filter(student=user)
    else:  # Staff, lecturers, deanery
        base_query = models.Q(assigned_to=user) | models.Q(viewers=user)

        if user.role == 1:  # Lecturer
            base_queryset = Request.objects.filter(
                models.Q(assigned_to=user) | models.Q(viewers=user)
            ).distinct()
        elif user.role in [2, 3]:  # Staff, deanery - all requests
            base_query |= models.Q()

        base_queryset = Request.objects.filter(base_query).distinct()
    
    filtered_queryset = filter_requests(request, base_queryset)
    
    # Statistics based on filtered data
    total_requests = filtered_queryset.count()
    pending_requests = filtered_queryset.filter(status=0).count()
    approved_requests = filtered_queryset.filter(status=1).count()
    rejected_requests = filtered_queryset.filter(status=2).count()
    
    context = {
        'total_requests': total_requests,
        'pending_requests': pending_requests,
        'approved_requests': approved_requests,
        'rejected_requests': rejected_requests,
        'total_users': User.objects.count(),
        'departments': Department.objects.all(),  # for filter dropdown
    }
    

    if user.role == 0:  # Student
        recent_requests = filtered_queryset.order_by('-created')[:5]
        status_counts = [pending_requests, approved_requests, rejected_requests]

        context.update({
            'recent_requests': recent_requests,
            'status_counts': status_counts,
        })
    else:  # Staff, lecturers, deanery
        # SLA statistics
        overdue_requests = [req for req in filtered_queryset.filter(status=0) if req.get_sla_status() == "בחריגה"]
        at_risk_requests = [req for req in filtered_queryset.filter(status=0) if req.get_sla_status() == "בסיכון"]

        # Recently assigned requests
        assigned_requests = filtered_queryset.filter(assigned_to=user, status=0).order_by('-created')[:5]

        # Recent activity
        recent_status_updates = RequestStatus.objects.filter(
            request__in=filtered_queryset).select_related('request', 'updated_by').order_by('-timestamp')[:10]

        # Last 30 days statistics
        recent_requests = filtered_queryset.filter(created__gte=thirty_days_ago)
        recent_resolved = recent_requests.filter(status__in=[1, 2])
        avg_resolution_time = 0

        if recent_resolved.exists():
            resolution_times = []
            for req in recent_resolved:
                if req.resolved_date:
                    delta = req.resolved_date - req.created
                    resolution_times.append(delta.total_seconds() / 3600)  # in hours

            if resolution_times:
                avg_resolution_time = sum(resolution_times) / len(resolution_times)

        if user.role in [2, 3]:  
            dept_counts = {}
            departments = Department.objects.all()
            for dept in departments:
                dept_counts[dept.name] = filtered_queryset.filter(dept=dept).count()
        else:
            dept_counts = None

        status_counts = [pending_requests, approved_requests, rejected_requests]

        # Pipeline status counts for chart
        pipeline_counts = []
        pipeline_statuses = dict(Request.PIPELINE_STATUSES)
        for status_value in sorted(pipeline_statuses.keys()):
            count = filtered_queryset.filter(pipeline_status=status_value).count()
            pipeline_counts.append({'status': pipeline_statuses[status_value], 'count': count})
            
        if user.role == 2 and user.department:  
            department = user.department
            

            this_month_start = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            last_month_start = (this_month_start - timezone.timedelta(days=1)).replace(day=1)
            
            dept_requests = Request.objects.filter(dept=department)
            this_month_requests = dept_requests.filter(created__gte=this_month_start).count()
            last_month_requests = dept_requests.filter(created__gte=last_month_start, created__lt=this_month_start).count()
            
            # חישוב אחוז השינוי
            if last_month_requests > 0:
                month_change_pct = ((this_month_requests - last_month_requests) / last_month_requests) * 100
            else:
                month_change_pct = 100 if this_month_requests > 0 else 0
            
            # ספירות משתמשים לפי סוג
            lecturer_count = User.objects.filter(department=department, role=1, is_active=True).count()
            student_count = User.objects.filter(department=department, role=0).count()
            course_count = Course.objects.filter(dept=department).count()
            
            # מרצים ממתינים לאישור
            pending_lecturers = User.objects.filter(role=1, is_active=False, department=department).count()
            
            # מידע על קורסים במחלקה
            courses = Course.objects.filter(dept=department)
            course_data = []
            for course in courses:
                course_lecturers = User.objects.filter(courses=course, role=1)
                student_count_course = User.objects.filter(courses=course, role=0).count()
                request_count = Request.objects.filter(course=course).count()
                
                course_data.append({
                    'course': course,
                    'lecturers': course_lecturers[:3],  # 3 המרצים הראשונים
                    'lecturer_count': course_lecturers.count(),
                    'student_count': student_count_course,
                    'request_count': request_count
                })
            
            # יצירת מילון נתונים מהמחלקה לתבנית
            stats = {
                'total_requests': total_requests,
                'this_month_requests': this_month_requests,
                'month_change_pct': round(month_change_pct, 1),
                'lecturer_count': lecturer_count,
                'student_count': student_count,
                'course_count': course_count,
                'overdue_count': len(overdue_requests),
                'at_risk_count': len(at_risk_requests),
            }
            
            # הוספת הנתונים לקונטקסט
            context.update({
                'stats': stats,
                'pending_count': pending_lecturers,
                'courses': courses,
                'course_data': course_data,
            })

        context.update({
            'assigned_requests': assigned_requests,
            'overdue_count': len(overdue_requests),
            'at_risk_count': len(at_risk_requests),
            'recent_status_updates': recent_status_updates,
            'avg_resolution_time': round(avg_resolution_time, 1),
            'dept_counts': dept_counts,
            'status_counts': status_counts,
            'pipeline_counts': pipeline_counts,
        })

    return render(request, 'dashboard.html', context)


@login_required
def export_dashboard_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    import io

    user = request.user

    if user.role == 0:  # Student
        base_queryset = Request.objects.filter(student=user)
    elif user.role in [1, 2, 3]:
        base_queryset = Request.objects.filter(
            Q(assigned_to=user) | Q(viewers=user)
        ).distinct()
        if user.role in [2, 3]:
            base_queryset = Request.objects.all()
    else:
        base_queryset = Request.objects.none()

    filtered_queryset = filter_requests(request, base_queryset)
    
    total_requests = filtered_queryset.count()
    pending_requests = filtered_queryset.filter(status=0).count()
    approved_requests = filtered_queryset.filter(status=1).count()
    rejected_requests = filtered_queryset.filter(status=2).count()

    overdue_count = len([r for r in filtered_queryset.filter(status=0) if r.get_sla_status() == "בחריגה"])
    at_risk_count = len([r for r in filtered_queryset.filter(status=0) if r.get_sla_status() == "בסיכון"])
    on_time_count = pending_requests - overdue_count - at_risk_count
    avg_resolution_time = 0

    recent_resolved = filtered_queryset.filter(status__in=[1, 2])
    if recent_resolved.exists():
        resolution_times = [
            (r.resolved_date - r.created).total_seconds() / 3600
            for r in recent_resolved if r.resolved_date
        ]
        if resolution_times:
            avg_resolution_time = round(sum(resolution_times) / len(resolution_times), 1)

    # יצירת קובץ Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard Stats"

    headers = ["מדד", "ערך"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    data = [
        ("סה\"כ בקשות", total_requests),
        ("בקשות בטיפול", pending_requests),
        ("בקשות שאושרו", approved_requests),
        ("בקשות שנדחו", rejected_requests),
        ("חורגות מזמן", overdue_count),
        ("בקשות בסיכון", at_risk_count),
        ("בקשות בזמן", on_time_count),
        ("זמן טיפול ממוצע (שעות)", avg_resolution_time),
    ]

    for row in data:
        ws.append(row)

    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 5

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=dashboard_stats.xlsx'
    return response


def filter_requests(request, queryset):
    """
    פונקציה לסינון בקשות בהתאם לפרמטרים בבקשת GET
    """
    # filter according to status
    status = request.GET.get('status')
    if status and status.isdigit():
        queryset = queryset.filter(status=int(status))

    # filter according to priority
    priority = request.GET.get('priority')
    if priority and priority.isdigit():
        queryset = queryset.filter(priority=int(priority))

    # filter according to department
    department = request.GET.get('department')
    if department and department.isdigit():
        queryset = queryset.filter(dept_id=int(department))

    # filter according to sla
    sla = request.GET.get('sla')
    if sla:
        pending = queryset.filter(status=0)
        if sla == 'overdue':
            overdue_ids = [req.id for req in pending if req.get_sla_status() == "בחריגה"]
            queryset = queryset.filter(id__in=overdue_ids)
        elif sla == 'at_risk':
            at_risk_ids = [req.id for req in pending if req.get_sla_status() == "בסיכון"]
            queryset = queryset.filter(id__in=at_risk_ids)
        elif sla == 'on_track':
            on_track_ids = [req.id for req in pending if req.get_sla_status() == "בזמן"]
            queryset = queryset.filter(id__in=on_track_ids)

    # filter according to data range
    date_from = request.GET.get('date_from')
    if date_from:
        try:
            from datetime import datetime
            date_from = datetime.strptime(date_from, '%Y-%m-%d')
            queryset = queryset.filter(created__gte=date_from)
        except ValueError:
            pass

    date_to = request.GET.get('date_to')
    if date_to:
        try:
            from datetime import datetime
            date_to = datetime.strptime(date_to, '%Y-%m-%d')
            date_to = date_to.replace(hour=23, minute=59, second=59)
            queryset = queryset.filter(created__lte=date_to)
        except ValueError:
            pass

    return queryset.order_by('-created')


def export_requests_csv(request):
    filtered_requests = get_filtered_requests(request)
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="Requests_export.xlsx"'
    writer = csv.writer(response)
    writer.writerow([
        'מס׳ בקשה', 'סוג בקשה', 'סטודנט', 'מחלקה', 'קורס', 'סטטוס',
        'שלב בתהליך', 'דחיפות', 'תאריך יצירה', 'תאריך טיפול', 'מטפלים'
    ])

    for req in filtered_requests:
        assigned_users = ", ".join([user.get_full_name() for user in req.assigned_to.all()])
        course_name = req.course.name if req.course else ""
        resolved_date = req.resolved_date.strftime('%d/%m/%Y %H:%M') if req.resolved_date else ""
        writer.writerow([
            req.id,
            req.get_title_display(),
            req.student.get_full_name(),
            req.dept.name,
            course_name,
            req.get_status_display(),
            req.get_current_status_display(),
            req.get_priority_display(),
            req.created.strftime('%d/%m/%Y %H:%M'),
            resolved_date,
            assigned_users
        ])

    return response

def assign_lecturer_courses_form(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    lecturers = User.objects.filter(role=1)

    if request.method == 'POST':
        selected_ids = request.POST.getlist('lecturers')
        course.user_set.set(User.objects.filter(id__in=selected_ids))
        return redirect('manage_courses')  

    context = {
        'course': course,
        'lecturers': lecturers,
    }
    return render(request, 'assign_lecturer_courses.html', context)

def export_requests_excel(request):
    filtered_requests = get_filtered_requests(request)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Requests"

    # Define headers
    headers = [
        'מס׳ בקשה', 'סוג בקשה', 'סטודנט', 'מחלקה', 'קורס', 'סטטוס',
        'שלב בתהליך', 'דחיפות', 'תאריך יצירה', 'תאריך טיפול', 'הערות', 'מטפלים'
    ]

    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Write data rows
    row_num = 2
    for req in filtered_requests:
        assigned_users = ", ".join([user.get_full_name() for user in req.assigned_to.all()])
        course_name = req.course.name if req.course else ""
        resolved_date = req.resolved_date.strftime('%d/%m/%Y %H:%M') if req.resolved_date else ""

        row = [
            req.id,
            req.get_title_display(),
            req.student.get_full_name(),
            req.dept.name,
            course_name,
            req.get_status_display(),
            req.get_current_status_display(),
            req.get_priority_display(),
            req.created.strftime('%d/%m/%Y %H:%M'),
            resolved_date,
            req.resolution_notes[:250] if req.resolution_notes else "",
            assigned_users
        ]

        for col_num, cell_value in enumerate(row, 1):
            ws.cell(row=row_num, column=col_num, value=cell_value)

        row_num += 1

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = 'attachment; filename="Requests_export.xlsx"'

    return response

def get_filtered_requests(request):
    user = request.user
    if user.role == 0:  # Student
        base_queryset = Request.objects.filter(student=user)
    elif user.role in [1, 2, 3]:  # Lecturer, Staff, Deanery
        base_queryset = Request.objects.filter(models.Q(viewers=user) | models.Q(assigned_to=user)).distinct()
    else:
        base_queryset = Request.objects.none()

    search_query = request.GET.get('q')
    if search_query:
        base_queryset = base_queryset.filter(models.Q(description__icontains=search_query) |
            models.Q(resolution_notes__icontains=search_query) | models.Q(student__first_name__icontains=search_query) |
            models.Q(student__last_name__icontains=search_query))

    status = request.GET.get('status')
    if status and status.isdigit():
        base_queryset = base_queryset.filter(status=int(status))

    pipeline_status = request.GET.get('pipeline_status')
    if pipeline_status and pipeline_status.isdigit():
        base_queryset = base_queryset.filter(pipeline_status=int(pipeline_status))

    priority = request.GET.get('priority')
    if priority and priority.isdigit():
        base_queryset = base_queryset.filter(priority=int(priority))

    department = request.GET.get('department')
    if department and department.isdigit():
        base_queryset = base_queryset.filter(dept_id=int(department))

    sla = request.GET.get('sla')
    if sla:
        pending = base_queryset.filter(status=0)
        if sla == 'overdue':
            overdue_ids = [req.id for req in pending if req.get_sla_status() == "Overdue"]
            base_queryset = base_queryset.filter(id__in=overdue_ids)
        elif sla == 'at_risk':
            at_risk_ids = [req.id for req in pending if req.get_sla_status() == "At Risk"]
            base_queryset = base_queryset.filter(id__in=at_risk_ids)
        elif sla == 'on_track':
            on_track_ids = [req.id for req in pending if req.get_sla_status() == "On Track"]
            base_queryset = base_queryset.filter(id__in=on_track_ids)

    date_from = request.GET.get('date_from')
    if date_from:
        try:
            from datetime import datetime
            date_from = datetime.strptime(date_from, '%Y-%m-%d')
            base_queryset = base_queryset.filter(created__gte=date_from)
        except ValueError:
            pass

    date_to = request.GET.get('date_to')
    if date_to:
        try:
            from datetime import datetime
            date_to = datetime.strptime(date_to, '%Y-%m-%d')
            date_to = date_to.replace(hour=23, minute=59, second=59)
            base_queryset = base_queryset.filter(created__lte=date_to)
        except ValueError:
            pass

    # newest first
    return base_queryset.order_by('-created')

def base(request):
    loginuser = request.user
    role = loginuser.role if loginuser.is_authenticated else 0

    # רק אם המשתמש ביקש באופן מפורש לראות התראות
    if request.GET.get('show_notifications') == '1':
        notifications = []
        notification_count = 0

        if loginuser.is_authenticated:
            # מזכירות (role == 2)
            if role == 2:
                new_requests = Request.objects.filter(status=0).count()
                pending_lecturers = User.objects.filter(role=1, is_active=False).count()

                if new_requests:
                    notifications.append(f"{new_requests} בקשות חדשות ממתינות לטיפול")
                    notification_count += new_requests

                if pending_lecturers:
                    notifications.append(f"{pending_lecturers} מרצים ממתינים לאישור")
                    notification_count += pending_lecturers

            # מרצה (role == 1)
            elif role == 1:
                if not loginuser.is_active:
                    notifications.append("החשבון שלך עדיין ממתין לאישור")
                    notification_count += 1
                else:
                    assigned_requests = Request.objects.filter(assigned_to=loginuser, status=0).count()
                    if assigned_requests:
                        notifications.append(f"{assigned_requests} בקשות חדשות הוקצו אליך")
                        notification_count += assigned_requests

            # דיקאנט (role == 3)
            elif role == 3:
                new_requests = Request.objects.filter(status=0).count()
                if new_requests:
                    notifications.append(f"{new_requests} בקשות חדשות ממתינות לבדיקה")
                    notification_count += new_requests

            # סטודנט (role == 0)
            elif role == 0:
                student_requests = Request.objects.filter(student=loginuser).order_by('-updated_at')[:5]
                for req in student_requests:
                    notifications.append(f"הבקשה #{req.id} - {req.get_status_display()}")
                    notification_count += 1
    else:
        # אם לא התבקשו במפורש התראות, החזר רשימה ריקה
        notifications = []
        notification_count = 0

    return render(request, 'base.html', {
        'role': role,
        'notifications': notifications,
        'notification_count': notification_count
    })




def approve_lecturer(request):
    if not request.user.is_authenticated or request.user.role != 2:
        messages.error(request, "אין לך הרשאות לבצע פעולה זו.")
        return redirect('home')

    if request.method == 'POST':
        action = request.POST.get('action')
        user_id = request.POST.get('user_id')

        try:
            lecturer = User.objects.get(id=user_id, role=1)
            if action == 'approve':
                lecturer.is_active = True
                lecturer.save()
                messages.success(request, f"המרצה {lecturer.get_full_name()} אושר בהצלחה.")
                send_lecturer_approval_email(lecturer, approved=True)

            elif action == 'reject':
                send_lecturer_approval_email(lecturer, approved=False)
                lecturer.delete() #change?
                messages.success(request, f"בקשת המרצה {lecturer.get_full_name()} נדחתה.")

        except User.DoesNotExist:
            messages.error(request, "המשתמש לא נמצא.")

    return redirect('dashboard')


from django.contrib import messages
from django.shortcuts import render, redirect
from .models import Course, User

def add_course(request):
    if not request.user.is_authenticated or request.user.role != 2:
        messages.error(request, "אין לך הרשאות לבצע פעולה זו.")
        return redirect('home')

    department = request.user.department
    if not department:
        messages.error(request, "לא משוייך למחלקה. אנא פנה למנהל המערכת.")
        return redirect('home')

    lecturers = User.objects.filter(role=1, is_active=True, department=department)

    if request.method == 'POST':
        name = request.POST.get('name')
        year = request.POST.get('year')
        lecturer_ids = request.POST.getlist('lecturers')

        # בדיקה שהשנה היא מספר שלם
        if not year.isdigit():
            messages.error(request, "שנת הלימוד חייבת להיות מספר.")
            return render(request, 'add_course.html', {
                'lecturers': lecturers,
                'name': name,
                'year': year,
                'selected_lecturers': User.objects.filter(id__in=lecturer_ids)
            })

        course = Course.objects.create(name=name, dept_id=department.id, year=int(year))

        if lecturer_ids:
            lecturers_to_assign = User.objects.filter(id__in=lecturer_ids)
            for lecturer in lecturers_to_assign:
                lecturer.courses.add(course)

        messages.success(request, f"הקורס '{name}' נוסף בהצלחה למערכת.")
        return redirect('manage_courses')  

    return render(request, 'add_course.html', {'lecturers': lecturers})

def student_requests(request, student_id):
    if not request.user.is_authenticated or request.user.role != 2:
        messages.error(request, "אין לך הרשאות לצפות בדף זה.")
        return redirect('home')

    try:
        student = User.objects.get(id=student_id, role=0)
    except User.DoesNotExist:
        messages.error(request, "הסטודנט לא נמצא.")
        return redirect('dashboard')

    if request.user.role == 1 and request.user.department != student.department:
        messages.error(request, "אין לך הרשאות לצפות בבקשות של סטודנט זה.")
        return redirect('dashboard')

    requests = Request.objects.filter(student=student).order_by('-created')
    return render(request, 'student_requests.html', {'student': student,'requests': requests})

def send_lecturer_approval_email(lecturer, approved=True):
    subject = "EasyReq - אישור חשבון מרצה" if approved else "EasyReq - דחיית בקשת רישום"
    context = {'lecturer': lecturer,'approved': approved,'login_url': settings.BASE_URL + '/login/'
    if hasattr(settings, 'BASE_URL') else 'http://localhost:8000/login/'}

    html_message = render_to_string('lecturer_approval_email.html', context)
    plain_message = "חשבונך במערכת EasyReq אושר. ניתן להתחבר באמצעות הקישור." if approved else "בקשת הרישום שלך למערכת EasyReq נדחתה."

    try:
        send_mail(subject,plain_message,settings.DEFAULT_FROM_EMAIL,[lecturer.email],html_message=html_message,
                  fail_silently=False)
        logger.debug(f"Approval email sent to lecturer {lecturer.email}")
    except Exception as e:
        logger.debug(f"Failed to send approval email: {e}")


def edit_student(request, student_id):
    if not request.user.is_authenticated or request.user.role != 2:
        messages.error(request, "אין לך הרשאות לבצע פעולה זו.")
        return redirect('home')

    try:
        student = User.objects.get(id=student_id, role=0)
    except User.DoesNotExist:
        messages.error(request, "הסטודנט לא נמצא.")
        return redirect('dashboard')

    if request.user.department != student.department:
        messages.error(request, "אין לך הרשאות לערוך פרטים של סטודנט זה.")
        return redirect('dashboard')

    if request.method == 'POST':
        student.first_name = request.POST.get('first_name')
        student.last_name = request.POST.get('last_name')
        student.email = request.POST.get('email')
        student.info = request.POST.get('info')
        student.is_active = 'is_active' in request.POST
        student.save()
        messages.success(request, f"פרטי הסטודנט {student.get_full_name()} עודכנו בהצלחה.")

    return redirect('student_requests', student_id=student_id)

@login_required
def manage_courses(request):
    if request.user.role != 2:
        return redirect('home')

    department = request.user.department
    courses = Course.objects.filter(dept=department).prefetch_related('user_set')

    for course in courses:
        all_users = list(course.user_set.all())
        course.related_lecturers = [u for u in all_users if u.role == 1]  
        course.student_count = sum(1 for u in all_users if u.role == 0)   

    return render(request, 'manage_courses.html', {'courses': courses})

def delete_course(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    course.delete()
    return redirect('manage_courses')
def edit_course(request, course_id):
    if not request.user.is_authenticated or request.user.role != 2:
        messages.error(request, "אין לך הרשאות לערוך קורסים.")
        return redirect('home')

    try:
        course = Course.objects.get(id=course_id)
    except Course.DoesNotExist:
        messages.error(request, "הקורס לא נמצא.")
        return redirect('manage_courses')

    lecturers = User.objects.filter(role=1, department=course.dept)

    if request.method == 'POST':
        course.name = request.POST.get('name')
        course.year = request.POST.get('year')
        selected_lecturers = request.POST.getlist('lecturers')
        course.save()

        # עדכון המרצים
        course.user_set.remove(*course.user_set.filter(role=1))
        course.user_set.add(*User.objects.filter(id__in=selected_lecturers, role=1))

        messages.success(request, f"הקורס {course.name} עודכן בהצלחה.")
        return redirect('manage_courses')

    return render(request, 'edit_course.html', {
        'course': course,
        'lecturers': lecturers,
        'selected_lecturers': course.user_set.filter(role=1)
    })

def assign_student_courses(request, student_id):
    if not request.user.is_authenticated or request.user.role != 2:
        messages.error(request, "אין לך הרשאות לבצע פעולה זו.")
        return redirect('home')

    try:
        student = User.objects.get(id=student_id, role=0)
    except User.DoesNotExist:
        messages.error(request, "הסטודנט לא נמצא.")
        return redirect('dashboard')

    if request.user.department != student.department:
        messages.error(request, "אין לך הרשאות לערוך פרטים של סטודנט זה.")
        return redirect('dashboard')

    if request.method == 'POST':
        course_ids = request.POST.getlist('courses')
        student.courses.clear()
        if course_ids:
            courses = Course.objects.filter(id__in=course_ids, dept=student.department)
            student.courses.add(*courses)

        messages.success(request, f"הקורסים של הסטודנט {student.get_full_name()} עודכנו בהצלחה.")

    return redirect('student_requests', student_id=student_id)

def deactivate_student(request, student_id):
    if not request.user.is_authenticated or request.user.role != 2:
        messages.error(request, "אין לך הרשאות לבצע פעולה זו.")
        return redirect('home')

    try:
        student = User.objects.get(id=student_id, role=0)
    except User.DoesNotExist:
        messages.error(request, "הסטודנט לא נמצא.")
        return redirect('dashboard')

    if request.user.department != student.department:
        messages.error(request, "אין לך הרשאות לבצע פעולה זו על סטודנט זה.")
        return redirect('dashboard')

    if request.method == 'POST':
        student.is_active = False
        student.save()
        messages.success(request, f"חשבון הסטודנט {student.get_full_name()} הושבת בהצלחה.")

    return redirect('dashboard')


def edit_student_form(request, student_id):
    if not request.user.is_authenticated or request.user.role != 2:
        messages.error(request, "אין לך הרשאות לבצע פעולה זו.")
        return redirect('home')

    try:
        student = User.objects.get(id=student_id, role=0)
    except User.DoesNotExist:
        messages.error(request, "הסטודנט לא נמצא.")
        return redirect('dashboard')

    # Check department access
    if request.user.department != student.department:
        messages.error(request, "אין לך הרשאות לערוך פרטים של סטודנט זה.")
        return redirect('dashboard')

    if request.method == 'POST':
        student.first_name = request.POST.get('first_name')
        student.last_name = request.POST.get('last_name')
        student.email = request.POST.get('email')
        student.info = request.POST.get('info')
        student.is_active = 'is_active' in request.POST
        student.save()
        messages.success(request, f"פרטי הסטודנט {student.get_full_name()} עודכנו בהצלחה.")
        return redirect('dashboard')

    context = {'student': student,}
    return render(request, 'edit_student.html', context)
def assign_lecturers_to_course(request, course_id):
    if not request.user.is_authenticated or request.user.role != 2:
        messages.error(request, "אין לך הרשאות לבצע פעולה זו.")
        return redirect('home')

    try:
        course = Course.objects.get(id=course_id)
    except Course.DoesNotExist:
        messages.error(request, "הקורס לא נמצא.")
        return redirect('manage_courses')

    lecturers = User.objects.filter(role=1, department=course.dept)

    if request.method == 'POST':
        selected_ids = request.POST.getlist('lecturers')
        # הסרה רק של מרצים, כדי לא לפגוע בסטודנטים משויכים
        course.user_set.remove(*course.user_set.filter(role=1))
        if selected_ids:
            selected_lecturers = User.objects.filter(id__in=selected_ids, role=1)
            course.user_set.add(*selected_lecturers)

        messages.success(request, "המרצים שויכו בהצלחה לקורס.")
        return redirect('manage_courses')

    context = {
        'course': course,
        'lecturers': lecturers,
        'assigned_lecturers': course.user_set.filter(role=1)
    }
    return render(request, 'assign_lecturers_to_course.html', context)

def assign_students_to_course(request, course_id):
    if not request.user.is_authenticated or request.user.role != 2:
        messages.error(request, "אין לך הרשאות לבצע פעולה זו.")
        return redirect('home')

    try:
        course = Course.objects.get(id=course_id)
    except Course.DoesNotExist:
        messages.error(request, "הקורס לא נמצא.")
        return redirect('manage_courses')

    students = User.objects.filter(role=0, department=course.dept)

    if request.method == 'POST':
        selected_ids = request.POST.getlist('students')
        course.user_set.clear()
        if selected_ids:
            selected_students = User.objects.filter(id__in=selected_ids, role=0)
            course.user_set.add(*selected_students)

        messages.success(request, "הסטודנטים שויכו בהצלחה לקורס.")
        return redirect('manage_courses')

    context = {
        'course': course,
        'students': students,
        'assigned_students': course.user_set.filter(role=0)
    }
    return render(request, 'assign_students_to_course.html', context)

def assign_student_courses_form(request, student_id):
    if not request.user.is_authenticated or request.user.role != 2:
        messages.error(request, "אין לך הרשאות לבצע פעולה זו.")
        return redirect('home')

    try:
        student = User.objects.get(id=student_id, role=0)
    except User.DoesNotExist:
        messages.error(request, "הסטודנט לא נמצא.")
        return redirect('dashboard')

    if request.user.department != student.department:
        messages.error(request, "אין לך הרשאות לערוך פרטים של סטודנט זה.")
        return redirect('dashboard')

    department_courses = Course.objects.filter(dept=student.department)

    if request.method == 'POST':
        course_ids = request.POST.getlist('courses')
        student.courses.clear()
        if course_ids:
            courses = Course.objects.filter(id__in=course_ids, dept=student.department)
            student.courses.add(*courses)

        messages.success(request, f"הקורסים של הסטודנט {student.get_full_name()} עודכנו בהצלחה.")
        return redirect('dashboard')

    context = {'student': student,'courses': department_courses,'assigned_courses': student.courses.all(),}
    return render(request, 'assign_student_courses.html', context)


def confirm_deactivate_student(request, student_id):
    if not request.user.is_authenticated or request.user.role != 2:
        messages.error(request, "אין לך הרשאות לבצע פעולה זו.")
        return redirect('home')

    try:
        student = User.objects.get(id=student_id, role=0)
    except User.DoesNotExist:
        messages.error(request, "הסטודנט לא נמצא.")
        return redirect('dashboard')

    if request.user.department != student.department:
        messages.error(request, "אין לך הרשאות לבצע פעולה זו על סטודנט זה.")
        return redirect('dashboard')

    if request.method == 'POST':
        student.is_active = False
        student.save()
        messages.success(request, f"חשבון הסטודנט {student.get_full_name()} הושבת בהצלחה.")
        return redirect('dashboard')

    context = {'student': student,}
    return render(request, 'confirm_deactivate_student.html', context)


@login_required
@require_http_methods(["GET", "POST"])
def manage_users(request):
    if request.user.role != 2:
        messages.error(request, "אין לך הרשאות לצפות בדף זה.")
        return redirect('home')

    department = request.user.department

    if request.method == 'POST':
        user_id = request.POST.get('user_id')
        action = request.POST.get('action')

        try:
            user = User.objects.get(id=user_id, department=department)
        except User.DoesNotExist:
            messages.error(request, "המשתמש לא נמצא או שאינו שייך למחלקה שלך.")
            return redirect('manage_users')

        if action == 'approve' and user.role == 1:
            user.is_active = True
            user.save()
            messages.success(request, f'המרצה {user.get_full_name()} אושר בהצלחה.')

        elif action == 'reject' and user.role == 1:
            user.delete()
            messages.success(request, f'המרצה {user.get_full_name()} נדחה ונמחק.')

        elif action == 'deactivate' and user.role == 1:
            user.delete()
            messages.success(request, f'המרצה {user.get_full_name()} הושבת ונמחק מהמערכת.')

        elif action == 'delete':
            user.delete()
            messages.success(request, f'המשתמש {user.get_full_name()} נמחק מהמערכת.')

        return redirect('manage_users')

    # הצגת המשתמשים למחלקה
    students = User.objects.filter(role=0, department=department).order_by('year', 'last_name')
    active_lecturers = User.objects.filter(role=1, department=department, is_active=True).order_by('last_name')
    pending_lecturers = User.objects.filter(role=1, department=department, is_active=False).order_by('last_name')

    context = {
        'students': students,
        'active_lecturers': active_lecturers,
        'pending_lecturers': pending_lecturers,
    }

    return render(request, 'manage_users.html', context)


def activate_student(request, student_id):
    if not request.user.is_authenticated or request.user.role not in [2, 3]:
        messages.error(request, "אין לך הרשאות לבצע פעולה זו.")
        return redirect('home')

    try:
        student = User.objects.get(id=student_id, role=0)
    except User.DoesNotExist:
        messages.error(request, "הסטודנט לא נמצא.")
        return redirect('dashboard')

    if request.user.department != student.department:
        messages.error(request, "אין לך הרשאות לבצע פעולה זו על סטודנט זה.")
        return redirect('dashboard')

    student.is_active = True
    student.save()
    messages.success(request, f"חשבון הסטודנט {student.get_full_name()} הופעל בהצלחה.")
    return redirect('dashboard')

@login_required
def create_request(request):
    TITLE_PRIORITY_MAPPING = {
        0: 1,  # ערעור על ציון → בינונית
        1: 1,  # בקשה למועד מיוחד → בינונית
        2: 0,  # שקלול עבודה בית בציון הסופי → נמוכה
        3: 0,  # דחיית הגשת עבודה → נמוכה
        4: 1,  # שחרור מחובת הרשמה → בינונית
        5: 2,  # בקשה לפטור מקורס → גבוהה
        6: 1,  # בקשה לפטור מעבודת הגשה → בינונית
        7: 1,  # בקשה לפטור מדרישת קדם → בינונית
        8: 2,  # חריגה מיוחדת - דיקאנט → גבוהה
    }

    if request.method == 'POST':
        student = request.user
        dept_id = student.department_id
        course_id = request.POST.get('course')
        title_id = request.POST.get('title')
        description = request.POST.get('description')

        try:
            priority = TITLE_PRIORITY_MAPPING.get(int(title_id), 1)  # ברירת מחדל: בינונית
        except Exception:
            priority = 1

        attachment = request.FILES.get('attachment')
        new_request = Request(
            student=student,
            dept_id=dept_id,
            title=int(title_id),
            description=description,
            priority=priority,
            attachments=attachment
        )

        if course_id:
            new_request.course_id = course_id

        new_request.save()
        assigned_users = new_request.auto_assign()
        RequestStatus.objects.create(
            request=new_request,
            status=0,
            updated_by=request.user,
            notes="Request submitted"
        )

        if len(assigned_users) > 0:
            send_request_notification(new_request, 'created')
            for user in assigned_users:
                send_request_notification(new_request, 'assigned', user)

        return redirect('request_detail', request_id=new_request.id)

    departments = Department.objects.all()
    courses = request.user.courses.all() if request.user.role == 0 else []

    return render(request, 'create_request.html', {
        'departments': departments,
        'courses': courses,
        'title_choices': Request.TITLES,
    })

def website_chat_response(request):
    """
    פונקציה מתאמת שמטפלת בבקשות צ'אט מהוידג'ט
    """
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            user_message = data.get('message', '')

            # קבלת תשובה מ-ChatGPT
            bot_response = get_openai_response(user_message)

            return JsonResponse({
                'response': bot_response
            })
        except Exception as e:
            logger.debug(f"שגיאה בטיפול בבקשה: {str(e)}")
            return JsonResponse({
                'response': f"אירעה שגיאה: {str(e)}"
            })
    return JsonResponse({'error': 'שיטה לא נתמכת'}, status=405)

def get_openai_response(message):
    """
    מקבל תשובה ממודל שפה של OpenAI
    """
    try:
        logger.debug("מתחבר ל-OpenAI API...")
        # יצירת קליינט OpenAI
        client = OpenAI(
            api_key="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")

        # בניית הפרומפט והקשר עם הגבלות נוקשות
        system_prompt = """
        אתה עוזר וירטואלי באתר בקשות סטודנטים, שמטרתו היחידה היא לספק מידע ועזרה בנושאים הקשורים לבקשות סטודנטים בלבד.

        אתה מוגבל אך ורק לנושאים הבאים:
        1. הגשת ערעור על ציון
        2. בקשה למועד מיוחד
        3. שקלול עבודה בית בציון הסופי
        4. דחיית הגשת עבודה
        5. שחרור מחובת הרשמה
        6. בקשה לפטור מקורס
        7. בקשה לפטור מעבודת הגשה
        8. בקשה לפטור מדרישת קדם
        9. בקשה לחריגה מיוחדת - דיקאנט
        10. מידע על סטטוס בקשות
        11. עזרה בניסוח בקשות אקדמיות
        12. מידע על תהליך הגשת בקשות

        חשוב מאוד: אינך מורשה לענות על שאלות שאינן קשורות לרשימת הנושאים הללו. 
        אם נשאלת שאלה שאינה קשורה לנושאים הללו (כמו מתכונים, תחביבים, חדשות, מזג אוויר, וכדומה):
        1. סרב בנימוס לענות
        2. הסבר שאתה מתמחה רק בנושאי בקשות סטודנטים
        3. הצע לדבר על אחד מהנושאים המותרים ברשימה

        עבור כל אחת מהבקשות המותרות, עליך להסביר:
        1. מהו הנתיב במערכת להגשת הבקשה (איפה צריך ללחוץ)
        2. למי הבקשה מוגשת (מרצה, יועץ, דיקנט וכו')
        3. איזה מסמכים נדרשים לרוב להגשת הבקשה
        4. מהו זמן הטיפול המשוער בבקשה

        כאשר המשתמש בוחר באפשרות "מידע על סטטוס בקשות", עליך להסביר את הנתיב במערכת לצפייה בסטטוס הבקשות.

        כאשר המשתמש בוחר באפשרות "עזרה בניסוח בקשות", עליך לשאול אותו איזה סוג בקשה הוא רוצה לנסח ואז לספק לו תבנית מפורטת לניסוח הבקשה המבוקשת.

        תשובותיך צריכות להיות מפורטות, מקצועיות אך ידידותיות, ולהכיל את כל המידע הדרוש לסטודנט כדי להבין איך להגיש את הבקשה ומה לצפות בהמשך התהליך.

        הערה חשובה: אתה לא יודע את הנתיבים המדויקים במערכת, למי בדיוק מוגשת כל בקשה, או את המסמכים הספציפיים הנדרשים - לכן הצג מידע כללי ומסוגנן שיכול להתאים למערכות בקשות סטודנטים טיפוסיות. הדגש שמדובר במידע כללי וכי הנתיבים המדויקים עשויים להשתנות בהתאם למערכת הספציפית של המוסד האקדמי.
        """

        logger.debug("שולח בקשה ל-OpenAI API...")
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": message}
            ],
            max_tokens=1000
        )

        logger.debug("התקבלה תשובה מ-OpenAI API")
        # החזרת התשובה
        return response.choices[0].message.content
    except Exception as e:
        logger.debug(f"שגיאה בחיבור ל-OpenAI: {str(e)}")
        return f"אירעה שגיאה בתקשורת עם מערכת ה-AI. נא לנסות שוב מאוחר יותר. (שגיאה: {str(e)})"


@require_http_methods(["GET"])
def rating_page(request):

    all_reviews = Review.objects.select_related('user').order_by('-created_at')
    total_reviews = all_reviews.count()

    avg_rating_data = all_reviews.aggregate(avg_rating=Avg('rating'))
    average_rating = avg_rating_data['avg_rating'] or 0

    rating_breakdown = []
    max_count = 0

    for star in range(5, 0, -1):
        count = all_reviews.filter(rating=star).count()
        max_count = max(max_count, count)
        rating_breakdown.append({
            'stars': star,
            'count': count,
            'percentage': 0
        })

    # Calculate percentages for visual bars
    for item in rating_breakdown:
        if max_count > 0:
            item['percentage'] = round((item['count'] / max_count * 100), 1)

    # Check if current user has already reviewed
    user_has_reviewed = False
    if request.user.is_authenticated:
        user_has_reviewed = Review.objects.filter(user=request.user).exists()

    # Paginate reviews (show 10 per page)
    paginator = Paginator(all_reviews, 10)
    page_number = request.GET.get('page', 1)
    reviews = paginator.get_page(page_number)

    # Check if there are more pages
    has_more_reviews = reviews.has_next()

    context = {
        'average_rating': average_rating,
        'total_reviews': total_reviews,
        'rating_breakdown': rating_breakdown,
        'reviews': reviews,
        'user_has_reviewed': user_has_reviewed,
        'has_more_reviews': has_more_reviews,
        'page_number': page_number,
    }

    return render(request, 'rating.html', context)


@login_required
@require_http_methods(["POST"])
def submit_review(request):

    try:
        existing_review = Review.objects.filter(user=request.user).first()
        if existing_review:
            messages.warning(request, "כבר הגשת ביקורת על האתר. תוכל לערוך אותה במקום להגיש חדשה.")
            return redirect('rating_page')

        rating = request.POST.get('rating')
        message = request.POST.get('message', '').strip()

        if not rating or not rating.isdigit() or int(rating) not in range(1, 6):
            messages.error(request, "אנא בחר דירוג תקין (1-5 כוכבים).")
            return redirect('rating_page')

        review = Review.objects.create(
            user=request.user,
            rating=int(rating),
            message=message if message else None,
            created_at=timezone.now()
        )

        star_text = "כוכב" if int(rating) == 1 else "כוכבים"
        messages.success(request, f"תודה על הביקורת! דירגת את האתר ב-{rating} {star_text}.")
        logger.info(f"New review submitted by user {request.user.id}: {rating} stars")

    except Exception as e:
        logger.error(f"Error submitting review: {str(e)}")
        messages.error(request, "אירעה שגיאה בשליחת הביקורת. אנא נסה שוב.")

    return redirect('rating_page')


@login_required
@require_http_methods(["GET", "POST"])
def edit_review(request):

    try:
        review = Review.objects.get(user=request.user)
    except Review.DoesNotExist:
        messages.error(request, "לא נמצאה ביקורת קיימת לעריכה.")
        return redirect('rating_page')

    if request.method == 'POST':
        rating = request.POST.get('rating')
        message = request.POST.get('message', '').strip()

        if not rating or not rating.isdigit() or int(rating) not in range(1, 6):
            messages.error(request, "אנא בחר דירוג תקין (1-5 כוכבים).")
            return render(request, 'edit_review.html', {'review': review})

        review.rating = int(rating)
        review.message = message if message else None
        review.updated_at = timezone.now()
        review.save()

        messages.success(request, "הביקורת עודכנה בהצלחה!")
        return redirect('rating_page')

    context = {'review': review,'is_edit': True,}
    return render(request, 'edit_review.html', context)


@login_required
@require_http_methods(["POST"])
def delete_review(request):

    try:
        review = Review.objects.get(user=request.user)
        review.delete()
        messages.success(request, "הביקורת נמחקה בהצלחה.")
        logger.info(f"Review deleted by user {request.user.id}")
    except Review.DoesNotExist:
        messages.error(request, "לא נמצאה ביקורת למחיקה.")
    except Exception as e:
        logger.error(f"Error deleting review: {str(e)}")
        messages.error(request, "אירעה שגיאה במחיקת הביקורת.")

    return redirect('rating_page')


@require_http_methods(["GET"])
def get_rating_stats(request):

    all_reviews = Review.objects.all()

    if not all_reviews.exists():
        return JsonResponse({'average_rating': 0,'total_reviews': 0, 'rating_distribution': [0, 0, 0, 0, 0]})

    avg_rating = all_reviews.aggregate(avg_rating=Avg('rating'))['avg_rating']
    total_reviews = all_reviews.count()

    distribution = []
    for star in range(1, 6):
        count = all_reviews.filter(rating=star).count()
        distribution.append(count)

    return JsonResponse({'average_rating': round(avg_rating, 2),'total_reviews': total_reviews,
        'rating_distribution': distribution, 'last_updated': timezone.now().isoformat()})
